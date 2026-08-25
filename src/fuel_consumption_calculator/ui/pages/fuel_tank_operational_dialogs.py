from __future__ import annotations

from datetime import datetime, timezone
from math import isfinite
from pathlib import Path

from PySide6.QtCore import QDateTime, Qt
from PySide6.QtWidgets import (
    QComboBox, QDateTimeEdit, QDialog, QDialogButtonBox, QDoubleSpinBox, QFormLayout,
    QHBoxLayout, QLabel, QLineEdit, QMessageBox, QPushButton, QTableWidget,
    QTableWidgetItem, QTextEdit, QVBoxLayout,
)

from fuel_consumption_calculator.domain.fuel_tank import MEASUREMENT_TYPES, FuelTank, TankCalibrationPoint
from fuel_consumption_calculator.services.fuel_tank_service import FuelTankService


def calibration_to_wide_rows(points: list[TankCalibrationPoint]) -> tuple[list[object], list[list[object]]]:
    trims = sorted({point.trim_m for point in points})
    grouped: dict[tuple[float | None, float | None], dict[float, float]] = {}
    for point in points:
        grouped.setdefault((point.sounding_cm, point.ullage_cm), {})[point.trim_m] = point.volume_m3
    rows = [[sounding, ullage, *[volumes.get(trim) for trim in trims]] for (sounding, ullage), volumes in sorted(grouped.items())]
    return ["Sounding cm", "Ullage cm", *trims], rows


def export_calibration_xlsx(path: Path, tank: FuelTank, points: list[TankCalibrationPoint]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    headers, rows = calibration_to_wide_rows(points)
    workbook = Workbook(); sheet = workbook.active; sheet.title = "Calibration"
    sheet["A1"] = f"Tank: {tank.name}"; sheet["A2"] = "Units: cm, m, m3"
    for column, header in enumerate(headers, 1): sheet.cell(4, column, header).font = Font(bold=True)
    for row_index, row in enumerate(rows, 5):
        for column, value in enumerate(row, 1): sheet.cell(row_index, column, value)
    for column in sheet.columns: sheet.column_dimensions[column[0].column_letter].width = 16
    workbook.save(path)


def import_calibration_xlsx(path: Path, tank_id: int) -> list[TankCalibrationPoint]:
    from openpyxl import load_workbook
    sheet = load_workbook(path, data_only=True).active
    header_row = next((row for row in sheet.iter_rows(values_only=True) if row and str(row[0] or "").strip().lower().startswith("sounding")), None)
    if header_row is None: raise ValueError("Workbook must contain a Sounding cm header row.")
    header_index = next(index for index, row in enumerate(sheet.iter_rows(values_only=True), 1) if row == header_row)
    headers = list(header_row); has_ullage = len(headers) > 1 and str(headers[1] or "").strip().lower().startswith("ullage")
    start = 2 if has_ullage else 1
    try: trims = [float(value) for value in headers[start:] if value is not None]
    except (TypeError, ValueError) as error: raise ValueError("Trim headers must be numeric.") from error
    if not trims: raise ValueError("Workbook must contain at least one trim column.")
    result: list[TankCalibrationPoint] = []
    for row in sheet.iter_rows(min_row=header_index + 1, values_only=True):
        if not row or row[0] is None: continue
        sounding = _finite_optional(row[0], "Sounding")
        ullage = _finite_optional(row[1], "Ullage") if has_ullage and len(row) > 1 and row[1] is not None else None
        for trim, volume in zip(trims, row[start:]):
            if volume is not None: result.append(TankCalibrationPoint(None, tank_id, sounding, ullage, _finite(trim, "Trim"), _finite(volume, "Volume")))
    if not result: raise ValueError("Workbook contains no calibration volume cells.")
    return result


def generate_calibration_points(tank_id: int, max_reading: float, reading_interval: float, max_forward: float, max_aft: float, trim_interval: float, even_volume: float, forward_volume: float, aft_volume: float) -> list[TankCalibrationPoint]:
    if min(max_reading, reading_interval, trim_interval, even_volume, forward_volume, aft_volume) <= 0 or max_forward < 0 or max_aft < 0: raise ValueError("Generator inputs must be positive; trim limits may be zero.")
    readings = _series(0, max_reading, reading_interval); trims = sorted({*[-x for x in _series(0, max_forward, trim_interval)], *_series(0, max_aft, trim_interval)})
    points = []
    for trim in trims:
        maximum = even_volume + ((abs(trim) / max_forward) * (forward_volume - even_volume) if trim < 0 and max_forward else (trim / max_aft) * (aft_volume - even_volume) if trim > 0 and max_aft else 0)
        for sounding in readings: points.append(TankCalibrationPoint(None, tank_id, sounding, max_reading - sounding, trim, maximum * sounding / max_reading))
    return points


class CalibrationDialog(QDialog):
    def __init__(self, service: FuelTankService, tank: FuelTank, parent=None) -> None:
        super().__init__(parent); self._service, self._tank = service, tank; self.setWindowTitle("Tank Calibration"); self.resize(760, 520)
        layout = QVBoxLayout(self); layout.addWidget(QLabel(f"{tank.name}  ·  Capacity {tank.capacity_m3:.2f} m³"))
        self.status = QLabel(); layout.addWidget(self.status)
        self.table = QTableWidget(0, 4); self.table.setHorizontalHeaderLabels(("Sounding cm", "Ullage cm", "Trim m", "Volume m³")); layout.addWidget(self.table)
        actions = QHBoxLayout()
        for text, callback in (("Add Row", self.add_row), ("Delete Selected Row", self.delete_rows), ("Import Excel", self.import_excel), ("Export Excel", self.export_excel), ("Generate Table", self.generate)):
            button = QPushButton(text); button.clicked.connect(callback); actions.addWidget(button)
        layout.addLayout(actions); buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel); buttons.accepted.connect(self.save); buttons.rejected.connect(self.reject); layout.addWidget(buttons)
        self.set_points(service.list_calibration_points(tank.id)); self._refresh_status()

    def set_points(self, points: list[TankCalibrationPoint]) -> None:
        self.table.setRowCount(0)
        for point in points: self._append(point.sounding_cm, point.ullage_cm, point.trim_m, point.volume_m3)

    def points(self) -> list[TankCalibrationPoint]:
        result=[]
        for row in range(self.table.rowCount()):
            values=[self.table.item(row, column).text().strip() if self.table.item(row, column) else "" for column in range(4)]
            sounding = _finite_optional(values[0], "Sounding") if values[0] else None; ullage = _finite_optional(values[1], "Ullage") if values[1] else None
            if sounding is None and ullage is None: raise ValueError("Every calibration row requires Sounding or Ullage.")
            result.append(TankCalibrationPoint(None, self._tank.id, sounding, ullage, _finite(values[2], "Trim"), _finite(values[3], "Volume")))
        return result

    def add_row(self) -> None: self._append(None, None, 0, 0)
    def delete_rows(self) -> None:
        for row in sorted({index.row() for index in self.table.selectedIndexes()}, reverse=True): self.table.removeRow(row)
    def _append(self, sounding, ullage, trim, volume) -> None:
        row=self.table.rowCount(); self.table.insertRow(row)
        for column, value in enumerate((sounding, ullage, trim, volume)): self.table.setItem(row, column, QTableWidgetItem("" if value is None else str(value)))
    def save(self) -> None:
        try: self._service.replace_calibration_points(self._tank.id, self.points())
        except ValueError as error: QMessageBox.warning(self, "Calibration not saved", str(error)); return
        self.accept()
    def import_excel(self) -> None:
        from PySide6.QtWidgets import QFileDialog
        path, _ = QFileDialog.getOpenFileName(self, "Import Calibration", filter="Excel Workbook (*.xlsx)")
        if path:
            try: self.set_points(import_calibration_xlsx(Path(path), self._tank.id)); self._refresh_status()
            except Exception as error: QMessageBox.warning(self, "Import failed", str(error))
    def export_excel(self) -> None:
        from PySide6.QtWidgets import QFileDialog
        path, _ = QFileDialog.getSaveFileName(self, "Export Calibration", "calibration.xlsx", "Excel Workbook (*.xlsx)")
        if path:
            try: export_calibration_xlsx(Path(path), self._tank, self.points())
            except Exception as error: QMessageBox.warning(self, "Export failed", str(error))
    def generate(self) -> None:
        dialog = GenerateCalibrationDialog(self._tank, self)
        if dialog.exec() == QDialog.DialogCode.Accepted and QMessageBox.question(self, "Generate Calibration", "Generated calibration is an approximation and must be checked against the approved vessel table before operational use. Generate a review table?") == QMessageBox.StandardButton.Yes:
            try: self.set_points(generate_calibration_points(self._tank.id, *dialog.values())); self._refresh_status()
            except ValueError as error: QMessageBox.warning(self, "Generation failed", str(error))
    def _refresh_status(self) -> None: self.status.setText("Calibration configured" if self.table.rowCount() else "No calibration points configured")


class GenerateCalibrationDialog(QDialog):
    def __init__(self, tank: FuelTank, parent=None) -> None:
        super().__init__(parent); self.setWindowTitle("Generate Calibration Table")
        layout=QVBoxLayout(self); form=QFormLayout(); self.inputs=[]
        fields=(("Max Sounding / Reference Height cm",100), ("Sounding Interval cm",10), ("Max Forward Trim m",2), ("Max Aft Trim m",2), ("Trim Interval m",1), ("Max Volume at Even Keel m³",tank.capacity_m3), ("Max Volume at Maximum Forward Trim m³",tank.capacity_m3), ("Max Volume at Maximum Aft Trim m³",tank.capacity_m3))
        for label, default in fields:
            widget=QDoubleSpinBox(); widget.setRange(0,100000); widget.setDecimals(3); widget.setValue(default); form.addRow(label,widget); self.inputs.append(widget)
        layout.addLayout(form); layout.addWidget(QLabel("Forward trim is negative; aft trim is positive. Generated values are an approximation for review.")); buttons=QDialogButtonBox(QDialogButtonBox.StandardButton.Ok|QDialogButtonBox.StandardButton.Cancel); buttons.accepted.connect(self.accept); buttons.rejected.connect(self.reject); layout.addWidget(buttons)
    def values(self) -> tuple[float,...]: return tuple(widget.value() for widget in self.inputs)


class LegacyUpdateTankROBDialog(QDialog):
    def __init__(self, service: FuelTankService, tank: FuelTank, parent=None) -> None:
        super().__init__(parent); self._service, self._tank=service,tank; self.setWindowTitle("Update Tank ROB"); layout=QVBoxLayout(self)
        points=service.list_calibration_points(tank.id); self.types=[kind for kind in MEASUREMENT_TYPES if any((p.sounding_cm if kind=="SOUNDING" else p.ullage_cm) is not None for p in points)]
        if not self.types:
            layout.addWidget(QLabel("No calibration table is configured for this tank.")); button=QPushButton("Open Calibration"); button.clicked.connect(self._open_calibration); layout.addWidget(button); return
        form=QFormLayout(); self.time=QDateTimeEdit(QDateTime.currentDateTimeUtc()); self.time.setTimeSpec(Qt.TimeSpec.UTC); self.type=QComboBox(); self.type.addItems(self.types); self.reading=QLineEdit(); self.trim=QLineEdit("0"); self.temperature=QLineEdit(); self.remarks=QTextEdit(); self.preview=QLabel("Enter reading and trim to calculate volume.")
        form.addRow("Observation Time UTC",self.time); form.addRow("Measurement Type",self.type); form.addRow("Reading cm",self.reading); form.addRow("Trim m",self.trim); form.addRow("Temperature °C",self.temperature); form.addRow("Remarks",self.remarks); layout.addLayout(form); layout.addWidget(self.preview); layout.addWidget(QLabel("Temperature is recorded for future density/mass calculation and does not affect volume in the current version.")); buttons=QDialogButtonBox(QDialogButtonBox.StandardButton.Save|QDialogButtonBox.StandardButton.Cancel); buttons.accepted.connect(self.save); buttons.rejected.connect(self.reject); layout.addWidget(buttons); self.reading.textChanged.connect(self.update_preview); self.trim.textChanged.connect(self.update_preview); self.type.currentTextChanged.connect(self.update_preview); self.update_preview()
    def _open_calibration(self) -> None: CalibrationDialog(self._service,self._tank,self).exec()
    def update_preview(self) -> None:
        try:
            volume=self._service.calculate_calibrated_volume(self._tank.id,self.type.currentText(),float(self.reading.text()),float(self.trim.text())); self.preview.setText(f"Calculated Volume: {volume:.3f} m³\nTank Fill: {volume/self._tank.capacity_m3*100:.1f}%"); self._valid=True
        except Exception as error: self.preview.setText(str(error)); self._valid=False
    def save(self) -> None:
        self.update_preview()
        if not self._valid: return
        try:
            temp=_finite_optional(self.temperature.text(),"Temperature") if self.temperature.text().strip() else None
            self._service.save_sounding_observation(tank_id=self._tank.id,reading_type=self.type.currentText(),reading_cm=float(self.reading.text()),trim_m=float(self.trim.text()),temperature_c=temp,fuel_batch_id=self._tank.current_fuel_batch_id,remarks=self.remarks.toPlainText().strip() or None,effective_at_utc=self.time.dateTime().toPython().replace(tzinfo=timezone.utc))
        except ValueError as error: QMessageBox.warning(self,"ROB not saved",str(error)); return
        self.accept()


class UpdateTankROBDialog(QDialog):
    def __init__(self, service: FuelTankService, tank: FuelTank, parent=None) -> None:
        super().__init__(parent)
        self._service, self._tank = service, tank
        self._batch = service.get_fuel_batch(tank.current_fuel_batch_id) if tank.current_fuel_batch_id else None
        self._snapshot = None
        self._valid = False
        self.setWindowTitle("Update Tank ROB")
        layout = QVBoxLayout(self)
        points = service.list_calibration_points(tank.id)
        self.types = [kind for kind in MEASUREMENT_TYPES if any((p.sounding_cm if kind == "SOUNDING" else p.ullage_cm) is not None for p in points)]
        if not self.types:
            layout.addWidget(QLabel("No calibration table is configured for this tank."))
            button = QPushButton("Open Calibration"); button.clicked.connect(self._open_calibration); layout.addWidget(button)
            return
        form = QFormLayout()
        self.time = QDateTimeEdit(QDateTime.currentDateTimeUtc()); self.time.setTimeSpec(Qt.TimeSpec.UTC)
        self.type = QComboBox(); self.type.addItems(self.types)
        self.reading = QLineEdit(); self.trim = QLineEdit("0"); self.temperature = QLineEdit(); self.manual_vcf = QLineEdit(); self.manual_vcf.setPlaceholderText("Optional, e.g. 0.98500")
        self.remarks = QTextEdit()
        form.addRow("Observation Time UTC", self.time); form.addRow("Measurement Type", self.type); form.addRow("Reading cm", self.reading); form.addRow("Trim m", self.trim); form.addRow("Temperature C", self.temperature)
        form.addRow("Fuel", QLabel(self._batch.fuel_type if self._batch else "UNKNOWN")); form.addRow("Batch", QLabel(self._batch.batch_name if self._batch else "No batch assigned")); form.addRow("Density @15°C", QLabel(f"{self._batch.density_15_kg_m3:.3f} kg/m³" if self._batch else "--")); form.addRow("Manual VCF", self.manual_vcf); form.addRow("Remarks", self.remarks)
        layout.addLayout(form)
        layout.addWidget(_muted_label("VCF is entered manually. Automatic ASTM/API calculation is not enabled."))
        self.preview = QLabel("Enter reading and trim to calculate volume."); self.preview.setWordWrap(True); layout.addWidget(self.preview)
        layout.addWidget(_muted_label("Temperature is recorded with the observation. VCF is entered manually in the current version."))
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self.save); buttons.rejected.connect(self.reject); layout.addWidget(buttons)
        self.reading.textChanged.connect(self.update_preview); self.trim.textChanged.connect(self.update_preview); self.type.currentTextChanged.connect(self.update_preview); self.manual_vcf.textChanged.connect(self.update_preview)
        self.update_preview()

    def _open_calibration(self) -> None:
        CalibrationDialog(self._service, self._tank, self).exec()

    def update_preview(self) -> None:
        self._snapshot = None
        reading_text = self.reading.text().strip()
        trim_text = self.trim.text().strip()
        if not reading_text or not trim_text:
            self.preview.setText("Enter reading and trim to calculate volume.")
            self._valid = False
            return
        try:
            reading = _finite(reading_text, "Reading")
            trim = _finite(trim_text, "Trim")
            volume = self._service.calculate_calibrated_volume(self._tank.id, self.type.currentText(), reading, trim)
            lines = [f"Observed Volume: {volume:.3f} m3", f"Fill: {volume / self._tank.capacity_m3 * 100:.1f}%"]
            vcf_text = self.manual_vcf.text().strip()
            if not vcf_text:
                lines.append("Mass snapshot: not recorded (Manual VCF not entered).")
            else:
                vcf = _finite(vcf_text, "Manual VCF")
                if self._batch is None:
                    lines.append("Mass snapshot: not recorded (no fuel batch assigned).")
                elif self._batch.density_15_kg_m3 < 100:
                    lines.append("Mass snapshot unavailable: batch density must be entered in kg/m3 (for example 978, not 0.978).")
                else:
                    result = self._service.calculate_manual_vcf_mass(volume, vcf, self._batch.density_15_kg_m3)
                    self._snapshot = (vcf, result.standard_volume_15_m3, self._batch.density_15_kg_m3, result.mass_mt)
                    lines.extend((f"Manual VCF: {vcf:.5f}", f"Volume @15 C: {result.standard_volume_15_m3:.3f} m3", f"Density @15 C: {self._batch.density_15_kg_m3:.3f} kg/m3", f"Calculated Mass: {result.mass_mt:.3f} MT"))
            self.preview.setText("\n".join(lines)); self._valid = True
        except ValueError as error:
            self.preview.setText(str(error)); self._valid = False

    def save(self) -> None:
        self.update_preview()
        if not self._valid:
            reading_text = self.reading.text().strip()
            trim_text = self.trim.text().strip()
            if not reading_text:
                QMessageBox.warning(self, "ROB not saved", "Reading is required.")
            elif not trim_text:
                QMessageBox.warning(self, "ROB not saved", "Trim is required.")
            else:
                QMessageBox.warning(self, "ROB not saved", self.preview.text())
            return
        try:
            temp = _finite_optional(self.temperature.text(), "Temperature") if self.temperature.text().strip() else None
            snapshot = self._snapshot
            self._service.save_sounding_observation(
                tank_id=self._tank.id, reading_type=self.type.currentText(), reading_cm=_finite(self.reading.text(), "Reading"), trim_m=_finite(self.trim.text(), "Trim"),
                temperature_c=temp, fuel_batch_id=self._tank.current_fuel_batch_id, remarks=self.remarks.toPlainText().strip() or None,
                effective_at_utc=self.time.dateTime().toPython().replace(tzinfo=timezone.utc),
                manual_vcf=snapshot[0] if snapshot else None, standard_volume_15_m3=snapshot[1] if snapshot else None,
                calculated_density_kg_m3=snapshot[2] if snapshot else None, calculated_mass_mt=snapshot[3] if snapshot else None,
            )
        except ValueError as error:
            QMessageBox.warning(self, "ROB not saved", str(error)); return
        self.accept()


def _muted_label(text: str) -> QLabel:
    label = QLabel(text); label.setWordWrap(True); label.setStyleSheet("color: #8caabd; font-size: 8pt;")
    return label


def _finite(value, label):
    try: value=float(value)
    except (TypeError,ValueError) as error: raise ValueError(f"{label} must be numeric.") from error
    if not isfinite(value): raise ValueError(f"{label} must be finite.")
    return value
def _finite_optional(value,label): return _finite(value,label)
def _series(start,end,step):
    values=[]; current=start
    while current<end: values.append(round(current,3)); current+=step
    return values+[round(end,3)]
